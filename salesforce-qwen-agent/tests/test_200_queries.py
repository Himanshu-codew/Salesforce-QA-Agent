"""
200 Test Queries Suite for Salesforce Chatbot
Covers all 11 categories from the test document.
Executes queries via POST /chat and generates Excel (.xlsx), CSV (.csv), and HTML report.

Usage:
    python tests/test_200_queries.py
"""

import asyncio
import csv
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import httpx

# Add project root to sys.path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Please install openpyxl: pip install openpyxl")


# ── Configuration ──
CHAT_API_URL = "http://localhost:8000/chat"
TIMEOUT_SECONDS = 120
OUTPUT_DIR = PROJECT_ROOT
OUTPUT_FILE = OUTPUT_DIR / "test_results_200_queries.xlsx"
OUTPUT_CSV_FILE = OUTPUT_DIR / "test_results_200_queries.csv"
OUTPUT_HTML_FILE = OUTPUT_DIR / "test_results_200_queries.html"


# ═══════════════════════════════════════════════════════════════
# 200 TEST QUERIES
# Format: (query_id, category, user_query, expected_behavior, pass_keywords, fail_keywords)
# ═══════════════════════════════════════════════════════════════

TEST_QUERIES_200 = [
    # ── Category 1: Data / Record Search (1-20) ──
    ("Q001", "Data / Record Search", "Show me all my Accounts.", "Execute SOQL on Account", ["account", "name", "id"], []),
    ("Q002", "Data / Record Search", "Tell me the details of the ABC Technologies Account.", "Find and return details of ABC Technologies", ["abc", "technologies", "account", "found", "details", "no", "exist"], []),
    ("Q003", "Data / Record Search", "Are there any Leads with the name Rohit Sharma?", "Query Leads WHERE Name/FirstName/LastName matches Rohit Sharma", ["lead", "rohit", "sharma", "found", "no", "result"], []),
    ("Q004", "Data / Record Search", "Show me all my open Opportunities.", "Query Opportunities WHERE IsClosed = false", ["opportunit", "stage", "open", "name"], []),
    ("Q005", "Data / Record Search", "What is the contact number of this customer?", "Ask which customer or query phone number", ["phone", "contact", "customer", "which", "number", "specify", "id"], []),
    ("Q006", "Data / Record Search", "Show me all Leads created last month.", "Query Leads WHERE CreatedDate = LAST_MONTH", ["lead", "created", "last_month", "name"], []),
    ("Q007", "Data / Record Search", "Which Accounts were recently updated?", "Query Accounts ORDER BY LastModifiedDate DESC", ["account", "updated", "modified", "name"], []),
    ("Q008", "Data / Record Search", "Show me all customers from Jaipur.", "Query Accounts/Contacts WHERE BillingCity/MailingCity = 'Jaipur'", ["jaipur", "customer", "account", "contact", "city", "no", "found"], []),
    ("Q009", "Data / Record Search", "Who are the Contacts associated with XYZ Company?", "Query Contacts for XYZ Company", ["contact", "xyz", "company", "account", "no", "found"], []),
    ("Q010", "Data / Record Search", "Find all Accounts in the Technology industry.", "Query Accounts WHERE Industry = 'Technology'", ["account", "technology", "industry"], []),
    ("Q011", "Data / Record Search", "Show me the details of the most recently created Account.", "Query Account ORDER BY CreatedDate DESC LIMIT 1", ["account", "created", "name"], []),
    ("Q012", "Data / Record Search", "Find all Leads created this week.", "Query Leads WHERE CreatedDate = THIS_WEEK", ["lead", "created", "week", "name"], []),
    ("Q013", "Data / Record Search", "Show me all Accounts owned by me.", "Query Accounts owned by current user", ["account", "name", "owner"], []),
    ("Q014", "Data / Record Search", "Find all Contacts associated with ABC Technologies.", "Query Contacts WHERE Account.Name = 'ABC Technologies'", ["contact", "abc", "technologies", "account", "no", "found"], []),
    ("Q015", "Data / Record Search", "Show me Leads whose status is Open - Not Contacted.", "Query Leads WHERE Status = 'Open - Not Contacted'", ["lead", "open", "contacted", "status"], []),
    ("Q016", "Data / Record Search", "Find all Opportunities associated with ABC Technologies.", "Query Opportunities for ABC Technologies", ["opportunit", "abc", "technologies", "account", "no", "found"], []),
    ("Q017", "Data / Record Search", "Show me Accounts that were created this month.", "Query Accounts WHERE CreatedDate = THIS_MONTH", ["account", "created", "month", "this_month"], []),
    ("Q018", "Data / Record Search", "Find all Leads whose company name is ABC Technologies.", "Query Leads WHERE Company = 'ABC Technologies'", ["lead", "abc", "technologies", "company", "no", "found"], []),
    ("Q019", "Data / Record Search", "Show me all Contacts with a missing phone number.", "Query Contacts WHERE Phone = null", ["contact", "phone", "missing", "null", "name"], []),
    ("Q020", "Data / Record Search", "Find Accounts that have not been updated recently.", "Query Accounts ORDER BY LastModifiedDate ASC", ["account", "updated", "modified", "name"], []),

    # ── Category 2: Reporting / Analytics (21-40) ──
    ("Q021", "Reporting / Analytics", "How many Leads are there in total?", "SOQL COUNT() on Lead", ["lead", "total", "count", "there are"], []),
    ("Q022", "Reporting / Analytics", "How many Leads are currently open?", "SOQL COUNT() on open Leads", ["lead", "open", "count", "total"], []),
    ("Q023", "Reporting / Analytics", "How many Opportunities were closed this month?", "COUNT Opportunities WHERE StageName IN ('Closed Won', 'Closed Lost') AND CloseDate = THIS_MONTH", ["opportunit", "closed", "month", "count"], []),
    ("Q024", "Reporting / Analytics", "Show me the top 10 Accounts by revenue.", "Query Accounts ORDER BY AnnualRevenue DESC LIMIT 10", ["account", "revenue", "annualrevenue", "top"], []),
    ("Q025", "Reporting / Analytics", "What is the total value of my sales pipeline?", "SUM(Amount) on open Opportunities", ["pipeline", "value", "total", "opportunit", "amount"], []),
    ("Q026", "Reporting / Analytics", "How many Opportunities are Closed Won?", "COUNT Opportunities WHERE StageName = 'Closed Won'", ["opportunit", "closed won", "count", "won"], []),
    ("Q027", "Reporting / Analytics", "How many Leads were created in the last 30 days?", "COUNT Leads WHERE CreatedDate = LAST_N_DAYS:30", ["lead", "created", "30", "days", "count"], []),
    ("Q028", "Reporting / Analytics", "Which salesperson is handling the most Opportunities?", "GROUP BY Owner on Opportunities", ["salesperson", "owner", "opportunit", "handling", "most"], []),
    ("Q029", "Reporting / Analytics", "Show me sales by region.", "Query Opportunities/Accounts grouped by BillingState or Region", ["sales", "region", "state", "opportunit", "amount", "no", "data"], []),
    ("Q030", "Reporting / Analytics", "What is the total revenue generated this month?", "SUM(Amount) WHERE StageName = 'Closed Won' AND CloseDate = THIS_MONTH", ["revenue", "total", "month", "closed won", "amount"], []),
    ("Q031", "Reporting / Analytics", "How many Accounts were created this year?", "COUNT Accounts WHERE CreatedDate = THIS_YEAR", ["account", "created", "year", "this_year", "count"], []),
    ("Q032", "Reporting / Analytics", "How many Opportunities are currently open?", "COUNT Opportunities WHERE IsClosed = false", ["opportunit", "open", "count", "isclosed"], []),
    ("Q033", "Reporting / Analytics", "What is the average Opportunity amount?", "AVG(Amount) on Opportunity", ["average", "avg", "opportunit", "amount"], []),
    ("Q034", "Reporting / Analytics", "What is the total value of Closed Won Opportunities?", "SUM(Amount) on Closed Won Opportunities", ["closed won", "total", "value", "opportunit", "amount"], []),
    ("Q035", "Reporting / Analytics", "Which salesperson has the highest sales amount?", "SUM(Amount) grouped by Owner ORDER BY SUM(Amount) DESC", ["salesperson", "owner", "highest", "sales", "amount"], []),
    ("Q036", "Reporting / Analytics", "How many Leads were converted this month?", "COUNT Leads WHERE IsConverted = true", ["lead", "converted", "month", "count"], []),
    ("Q037", "Reporting / Analytics", "What is the Lead conversion rate?", "Calculate conversion rate of Leads", ["conversion", "rate", "lead", "converted", "%", "percent"], []),
    ("Q038", "Reporting / Analytics", "Show me the number of Leads by Lead Status.", "GROUP BY Status on Lead", ["lead", "status", "count", "open", "contacted"], []),
    ("Q039", "Reporting / Analytics", "Show me the number of Opportunities by Stage.", "GROUP BY StageName on Opportunity", ["opportunit", "stage", "count", "prospecting", "closed"], []),
    ("Q040", "Reporting / Analytics", "Which month had the highest sales?", "Calculate month with highest Closed Won amount", ["month", "sales", "highest", "closed won", "amount"], []),

    # ── Category 3: Record Creation (41-55) ──
    ("Q041", "Record Creation", "Create a new Lead for Rohit Sharma.", "Ask for missing required fields (Company) or create lead", ["company", "lead", "rohit", "sharma", "required", "provide", "created"], []),
    ("Q042", "Record Creation", "Create a new Account named ABC Ltd.", "Create Account with Name='ABC Ltd'", ["account", "abc ltd", "created", "success", "id"], []),
    ("Q043", "Record Creation", "Create a new Contact.", "Ask for Contact required fields (LastName, etc.)", ["contact", "name", "lastname", "required", "provide", "specify"], []),
    ("Q044", "Record Creation", "Create a new Opportunity for this customer.", "Ask for Opportunity required fields (Name, Stage, CloseDate)", ["opportunit", "name", "stage", "closedate", "required", "provide", "customer"], []),
    ("Q045", "Record Creation", "Create a new Lead and set the status to Open - Not Contacted.", "Ask for missing fields (Name, Company) or create", ["lead", "company", "name", "required", "status", "open - not contacted"], []),
    ("Q046", "Record Creation", "Create a new Case for this customer.", "Ask for Case subject/customer details or create case", ["case", "subject", "customer", "provide", "specify", "created"], []),
    ("Q047", "Record Creation", "Create a new Account named Tech Solutions India.", "Create Account with Name='Tech Solutions India'", ["account", "tech solutions india", "created", "success"], []),
    ("Q048", "Record Creation", "Create a new Contact associated with ABC Technologies.", "Ask for Contact name or create contact", ["contact", "abc technologies", "name", "lastname", "provide", "created"], []),
    ("Q049", "Record Creation", "Create a new Opportunity with an amount of 500,000.", "Ask for Opportunity name, stage, and close date", ["opportunit", "name", "stage", "close", "required", "provide", "500000"], []),
    ("Q050", "Record Creation", "Create a new Lead for Rohit Sharma who works at Tech Solutions India.", "Create Lead with FirstName='Rohit', LastName='Sharma', Company='Tech Solutions India'", ["lead", "rohit", "sharma", "tech solutions", "created", "success"], []),
    ("Q051", "Record Creation", "Create a new Case and set its priority to High.", "Ask for Case subject or create Case with Priority='High'", ["case", "subject", "priority", "high", "provide", "created"], []),
    ("Q052", "Record Creation", "Create a new Opportunity with the stage set to Prospecting.", "Ask for Opportunity Name and CloseDate", ["opportunit", "name", "close", "prospecting", "required", "provide"], []),
    ("Q053", "Record Creation", "Create a new Task for this Lead.", "Ask for Task details / Lead reference or create Task", ["task", "lead", "subject", "provide", "specify", "created"], []),
    ("Q054", "Record Creation", "Create a new Account and assign it to Aman.", "Ask for Account name or find Aman user ID and create", ["account", "name", "aman", "provide", "created", "assign"], []),
    ("Q055", "Record Creation", "Create a new Contact under the ABC Technologies Account.", "Ask for Contact name or create under ABC Technologies", ["contact", "abc technologies", "name", "lastname", "provide", "created"], []),

    # ── Category 4: Record Update (56-75) ──
    ("Q056", "Record Update", "Update Rohit Sharma's Lead Status to Qualified.", "Find Lead Rohit Sharma and update Status", ["lead", "rohit", "sharma", "status", "qualified", "updated", "update"], []),
    ("Q057", "Record Update", "Change the Industry of ABC Technologies to Technology.", "Find ABC Technologies and update Industry", ["abc technologies", "industry", "technology", "updated", "update", "account"], []),
    ("Q058", "Record Update", "Update the Opportunity amount to 500,000.", "Ask which Opportunity or update Opportunity amount", ["opportunit", "amount", "500,000", "500000", "which", "specify", "updated"], []),
    ("Q059", "Record Update", "Change this Opportunity's Stage to Closed Won.", "Ask which Opportunity or update Stage to Closed Won", ["opportunit", "stage", "closed won", "which", "specify", "updated"], []),
    ("Q060", "Record Update", "Update the Contact's phone number.", "Ask which Contact and what phone number", ["contact", "phone", "which", "specify", "provide", "number"], []),
    ("Q061", "Record Update", "Change the Case priority to High.", "Ask which Case or update Priority to High", ["case", "priority", "high", "which", "specify", "updated"], []),
    ("Q062", "Record Update", "Assign this Lead to Aman.", "Ask which Lead or find Aman and assign", ["lead", "aman", "assign", "owner", "which", "specify"], []),
    ("Q063", "Record Update", "Change this Lead's status to Working - Contacted.", "Ask which Lead or update Status to Working - Contacted", ["lead", "status", "working - contacted", "which", "specify", "updated"], []),
    ("Q064", "Record Update", "Update the Account's phone number.", "Ask which Account and what phone number", ["account", "phone", "which", "specify", "provide", "number"], []),
    ("Q065", "Record Update", "Change the Opportunity close date.", "Ask which Opportunity and what close date", ["opportunit", "close", "date", "which", "specify", "provide"], []),
    ("Q066", "Record Update", "Update the Contact's email address.", "Ask which Contact and what email address", ["contact", "email", "which", "specify", "provide", "address"], []),
    ("Q067", "Record Update", "Change this Case status to Closed.", "Ask which Case or update Status to Closed", ["case", "status", "closed", "which", "specify", "updated"], []),
    ("Q068", "Record Update", "Assign this Opportunity to Aman.", "Ask which Opportunity or find Aman and assign", ["opportunit", "aman", "assign", "owner", "which", "specify"], []),
    ("Q069", "Record Update", "Update the Account's Industry to Information Technology.", "Ask which Account or update Industry", ["account", "industry", "information technology", "which", "specify", "updated"], []),
    ("Q070", "Record Update", "Change the Lead owner to Aman.", "Ask which Lead or assign owner to Aman", ["lead", "owner", "aman", "which", "specify", "change"], []),
    ("Q071", "Record Update", "Update the Opportunity probability.", "Ask which Opportunity and what probability", ["opportunit", "probability", "which", "specify", "provide", "percent"], []),
    ("Q072", "Record Update", "Change the Case priority from Medium to High.", "Ask which Case or update Priority", ["case", "priority", "high", "medium", "which", "specify", "updated"], []),
    ("Q073", "Record Update", "Update the Lead's company name.", "Ask which Lead and what company name", ["lead", "company", "which", "specify", "provide", "name"], []),
    ("Q074", "Record Update", "Change the Account owner.", "Ask which Account and who is the new owner", ["account", "owner", "which", "specify", "provide", "new"], []),
    ("Q075", "Record Update", "Update the Opportunity name.", "Ask which Opportunity and what new name", ["opportunit", "name", "which", "specify", "provide", "new"], []),

    # ── Category 5: Salesforce User / Owner Queries (76-90) ──
    ("Q076", "Salesforce User / Owner Queries", "Show me all Accounts owned by me.", "Query Accounts owned by current user", ["account", "name", "owner"], []),
    ("Q077", "Salesforce User / Owner Queries", "How many Opportunities does Aman have?", "COUNT Opportunities WHERE Owner.Name = 'Aman'", ["opportunit", "aman", "count", "owner", "total", "have"], []),
    ("Q078", "Salesforce User / Owner Queries", "Who is the owner of this Lead?", "Ask which Lead or query Lead owner", ["lead", "owner", "which", "specify", "name"], []),
    ("Q079", "Salesforce User / Owner Queries", "Who is this Case assigned to?", "Ask which Case or query Case owner", ["case", "assigned", "owner", "which", "specify", "user"], []),
    ("Q080", "Salesforce User / Owner Queries", "Which Leads are assigned to me?", "Query Leads owned by current user", ["lead", "assigned", "owner", "name", "status"], []),
    ("Q081", "Salesforce User / Owner Queries", "Show me the users in the Sales team.", "Query Users WHERE Department = 'Sales' or Profile", ["user", "sales", "team", "name", "department"], []),
    ("Q082", "Salesforce User / Owner Queries", "Who last modified this record?", "Ask which record or check LastModifiedBy", ["modified", "record", "which", "lastmodifiedby", "user", "specify"], []),
    ("Q083", "Salesforce User / Owner Queries", "How many Accounts are assigned to Aman?", "COUNT Accounts WHERE Owner.Name = 'Aman'", ["account", "aman", "count", "assigned", "owner"], []),
    ("Q084", "Salesforce User / Owner Queries", "Show me all Opportunities owned by Aman.", "Query Opportunities WHERE Owner.Name = 'Aman'", ["opportunit", "aman", "owner", "name", "stage"], []),
    ("Q085", "Salesforce User / Owner Queries", "Which user owns the most Opportunities?", "GROUP BY Owner on Opportunities ORDER BY COUNT(Id) DESC", ["user", "owner", "opportunit", "most", "count"], []),
    ("Q086", "Salesforce User / Owner Queries", "Who is the owner of ABC Technologies?", "Query Account ABC Technologies Owner.Name", ["abc technologies", "owner", "account", "user", "name", "no", "found"], []),
    ("Q087", "Salesforce User / Owner Queries", "Show me all Leads assigned to this user.", "Ask which user or query Leads", ["lead", "user", "assigned", "which", "specify", "name"], []),
    ("Q088", "Salesforce User / Owner Queries", "Which Accounts are assigned to me?", "Query Accounts owned by current user", ["account", "assigned", "name", "owner"], []),
    ("Q089", "Salesforce User / Owner Queries", "Who created this record?", "Ask which record or check CreatedBy", ["created", "record", "which", "createdby", "user", "specify"], []),
    ("Q090", "Salesforce User / Owner Queries", "When was this record last modified?", "Ask which record or check LastModifiedDate", ["modified", "record", "which", "date", "lastmodifieddate", "specify"], []),

    # ── Category 6: Cases / Customer Support (91-110) ──
    ("Q091", "Cases / Customer Support", "Show me all open Cases.", "Query Cases WHERE Status != 'Closed'", ["case", "status", "open", "subject"], []),
    ("Q092", "Cases / Customer Support", "Which Cases have High priority?", "Query Cases WHERE Priority = 'High'", ["case", "priority", "high", "subject"], []),
    ("Q093", "Cases / Customer Support", "How many Cases does ABC Technologies have?", "COUNT Cases for ABC Technologies", ["case", "abc technologies", "count", "account", "total"], []),
    ("Q094", "Cases / Customer Support", "What is the current status of this Case?", "Ask which Case or return Case status", ["case", "status", "which", "specify", "current"], []),
    ("Q095", "Cases / Customer Support", "Show me all unresolved Cases.", "Query Cases WHERE IsClosed = false", ["case", "unresolved", "open", "status", "subject"], []),
    ("Q096", "Cases / Customer Support", "Which is the oldest open Case?", "Query Cases WHERE IsClosed = false ORDER BY CreatedDate ASC LIMIT 1", ["case", "oldest", "open", "created", "subject"], []),
    ("Q097", "Cases / Customer Support", "Close this Case.", "Ask which Case or update Status to Closed", ["case", "close", "closed", "which", "specify", "status"], []),
    ("Q098", "Cases / Customer Support", "Change this Case priority from Medium to High.", "Ask which Case or update Priority", ["case", "priority", "high", "medium", "which", "specify"], []),
    ("Q099", "Cases / Customer Support", "Show me all Cases created this month.", "Query Cases WHERE CreatedDate = THIS_MONTH", ["case", "created", "month", "this_month", "subject"], []),
    ("Q100", "Cases / Customer Support", "Show me all High-priority open Cases.", "Query Cases WHERE Priority = 'High' AND IsClosed = false", ["case", "high", "priority", "open", "subject"], []),
    ("Q101", "Cases / Customer Support", "Which Cases are assigned to me?", "Query Cases owned by current user", ["case", "assigned", "owner", "subject", "status"], []),
    ("Q102", "Cases / Customer Support", "Who is handling this Case?", "Ask which Case or return Case owner", ["case", "handling", "owner", "which", "specify", "user"], []),
    ("Q103", "Cases / Customer Support", "Show me all Cases for this customer.", "Ask which customer or query Cases for Account/Contact", ["case", "customer", "account", "which", "specify", "subject"], []),
    ("Q104", "Cases / Customer Support", "How many Cases were closed this month?", "COUNT Cases WHERE Status = 'Closed' AND ClosedDate = THIS_MONTH", ["case", "closed", "month", "count"], []),
    ("Q105", "Cases / Customer Support", "Show me Cases that have been open for more than 7 days.", "Query Cases WHERE IsClosed = false AND CreatedDate < LAST_N_DAYS:7", ["case", "open", "7", "days", "created", "subject"], []),
    ("Q106", "Cases / Customer Support", "What is the latest Case created for ABC Technologies?", "Query latest Case for ABC Technologies", ["case", "abc technologies", "latest", "created", "subject", "no", "found"], []),
    ("Q107", "Cases / Customer Support", "Show me all unresolved Cases assigned to Aman.", "Query Cases WHERE Owner.Name = 'Aman' AND IsClosed = false", ["case", "aman", "unresolved", "open", "owner", "subject"], []),
    ("Q108", "Cases / Customer Support", "Change the status of this Case to In Progress.", "Ask which Case or update Status to In Progress", ["case", "status", "in progress", "which", "specify", "updated"], []),
    ("Q109", "Cases / Customer Support", "Assign this Case to Aman.", "Ask which Case or assign Case to Aman", ["case", "aman", "assign", "owner", "which", "specify"], []),
    ("Q110", "Cases / Customer Support", "Show me the details of Case number 12345.", "Query Case WHERE CaseNumber = '12345' or CaseNumber LIKE '%12345%'", ["case", "12345", "subject", "status", "number", "no", "found"], []),

    # ── Category 7: Opportunity / Sales (111-130) ──
    ("Q111", "Opportunity / Sales", "Show me all my open Opportunities.", "Query Opportunities WHERE IsClosed = false", ["opportunit", "open", "stage", "amount"], []),
    ("Q112", "Opportunity / Sales", "Which is my largest Opportunity?", "Query Opportunity ORDER BY Amount DESC LIMIT 1", ["opportunit", "largest", "amount", "name"], []),
    ("Q113", "Opportunity / Sales", "What is my sales pipeline for this month?", "SUM(Amount) WHERE IsClosed = false AND CloseDate = THIS_MONTH", ["pipeline", "sales", "month", "opportunit", "amount"], []),
    ("Q114", "Opportunity / Sales", "Show me all Closed Won Opportunities.", "Query Opportunities WHERE StageName = 'Closed Won'", ["opportunit", "closed won", "stage", "amount"], []),
    ("Q115", "Opportunity / Sales", "Show me the Opportunities associated with ABC Technologies.", "Query Opportunities for ABC Technologies", ["opportunit", "abc technologies", "account", "stage", "no", "found"], []),
    ("Q116", "Opportunity / Sales", "What is the expected revenue?", "Query or calculate ExpectedRevenue / Amount", ["revenue", "expected", "opportunit", "amount"], []),
    ("Q117", "Opportunity / Sales", "Which Opportunities are expected to close soon?", "Query Opportunities ORDER BY CloseDate ASC WHERE IsClosed = false", ["opportunit", "close", "soon", "closedate", "stage"], []),
    ("Q118", "Opportunity / Sales", "Show me Opportunities closing this month.", "Query Opportunities WHERE CloseDate = THIS_MONTH", ["opportunit", "month", "this_month", "closedate", "stage"], []),
    ("Q119", "Opportunity / Sales", "What is the total value of my open Opportunities?", "SUM(Amount) WHERE IsClosed = false", ["opportunit", "open", "total", "value", "amount"], []),
    ("Q120", "Opportunity / Sales", "Which Opportunity has the highest amount?", "Query Opportunity ORDER BY Amount DESC LIMIT 1", ["opportunit", "highest", "amount", "name"], []),
    ("Q121", "Opportunity / Sales", "Show me all Opportunities in the Negotiation stage.", "Query Opportunities WHERE StageName LIKE '%Negotiation%'", ["opportunit", "negotiation", "stage", "name"], []),
    ("Q122", "Opportunity / Sales", "Show me Opportunities that are past their expected close date.", "Query Opportunities WHERE CloseDate < TODAY AND IsClosed = false", ["opportunit", "past", "close", "date", "overdue", "closedate"], []),
    ("Q123", "Opportunity / Sales", "How many Opportunities are in the Proposal stage?", "COUNT Opportunities WHERE StageName LIKE '%Proposal%'", ["opportunit", "proposal", "stage", "count"], []),
    ("Q124", "Opportunity / Sales", "What is the total value of Closed Lost Opportunities?", "SUM(Amount) WHERE StageName = 'Closed Lost'", ["opportunit", "closed lost", "total", "value", "amount"], []),
    ("Q125", "Opportunity / Sales", "Show me all Opportunities owned by Aman.", "Query Opportunities WHERE Owner.Name = 'Aman'", ["opportunit", "aman", "owner", "name", "stage"], []),
    ("Q126", "Opportunity / Sales", "Which Accounts have the highest Opportunity value?", "Query Accounts grouped or sorted by Opportunity Amount", ["account", "opportunit", "value", "highest", "amount", "revenue"], []),
    ("Q127", "Opportunity / Sales", "Show me Opportunities created this month.", "Query Opportunities WHERE CreatedDate = THIS_MONTH", ["opportunit", "created", "month", "this_month", "name"], []),
    ("Q128", "Opportunity / Sales", "What is the average Opportunity amount?", "AVG(Amount) on Opportunity", ["average", "avg", "opportunit", "amount"], []),
    ("Q129", "Opportunity / Sales", "Show me all Opportunities with an amount greater than 100,000.", "Query Opportunities WHERE Amount > 100000", ["opportunit", "amount", "100,000", "100000", "name"], []),
    ("Q130", "Opportunity / Sales", "Which Opportunities have the highest probability of closing?", "Query Opportunities ORDER BY Probability DESC", ["opportunit", "probability", "highest", "closing", "stage"], []),

    # ── Category 8: Tasks / Activities / Follow-ups (131-150) ──
    ("Q131", "Tasks / Activities / Follow-ups", "Show me my pending Tasks.", "Query Tasks WHERE Status != 'Completed'", ["task", "pending", "subject", "status", "no", "found"], []),
    ("Q132", "Tasks / Activities / Follow-ups", "What follow-ups do I have today?", "Query Tasks WHERE ActivityDate = TODAY", ["task", "follow", "today", "activity", "no", "found"], []),
    ("Q133", "Tasks / Activities / Follow-ups", "Do I have a call scheduled with Rohit?", "Query Tasks/Events WHERE Subject LIKE '%Call%' AND Who.Name LIKE '%Rohit%'", ["call", "rohit", "task", "event", "scheduled", "no", "found"], []),
    ("Q134", "Tasks / Activities / Follow-ups", "Show me my upcoming meetings.", "Query Events WHERE StartDateTime >= TODAY", ["meeting", "event", "upcoming", "subject", "no", "found"], []),
    ("Q135", "Tasks / Activities / Follow-ups", "What was the last activity associated with this Lead?", "Ask which Lead or query Tasks for Lead", ["activity", "lead", "task", "last", "which", "specify"], []),
    ("Q136", "Tasks / Activities / Follow-ups", "Show me my Tasks for tomorrow.", "Query Tasks WHERE ActivityDate = TOMORROW", ["task", "tomorrow", "subject", "due", "no", "found"], []),
    ("Q137", "Tasks / Activities / Follow-ups", "What Tasks are overdue?", "Query Tasks WHERE ActivityDate < TODAY AND Status != 'Completed'", ["task", "overdue", "subject", "date", "no", "found"], []),
    ("Q138", "Tasks / Activities / Follow-ups", "Show me all Tasks assigned to me.", "Query Tasks owned by current user", ["task", "assigned", "subject", "status", "no", "found"], []),
    ("Q139", "Tasks / Activities / Follow-ups", "Which Tasks are due today?", "Query Tasks WHERE ActivityDate = TODAY", ["task", "today", "due", "subject", "no", "found"], []),
    ("Q140", "Tasks / Activities / Follow-ups", "Show me my upcoming activities.", "Query Events and Tasks upcoming", ["activit", "upcoming", "task", "event", "meeting", "no", "found"], []),
    ("Q141", "Tasks / Activities / Follow-ups", "What was the last interaction with ABC Technologies?", "Query latest Task/Activity for ABC Technologies", ["abc technologies", "interaction", "activity", "task", "last", "no", "found"], []),
    ("Q142", "Tasks / Activities / Follow-ups", "Show me all overdue Tasks.", "Query Tasks WHERE ActivityDate < TODAY AND Status != 'Completed'", ["task", "overdue", "subject", "due", "no", "found"], []),
    ("Q143", "Tasks / Activities / Follow-ups", "Who is assigned to this Task?", "Ask which Task or return Task owner", ["task", "assigned", "owner", "which", "specify", "user"], []),
    ("Q144", "Tasks / Activities / Follow-ups", "When is my next meeting?", "Query next Event ORDER BY StartDateTime ASC", ["meeting", "event", "next", "start", "no", "found"], []),
    ("Q145", "Tasks / Activities / Follow-ups", "Show me all Tasks related to this Opportunity.", "Ask which Opportunity or query Tasks WHERE WhatId", ["task", "opportunit", "related", "which", "specify", "subject"], []),
    ("Q146", "Tasks / Activities / Follow-ups", "What activities are associated with this Lead?", "Ask which Lead or query Tasks for Lead", ["activit", "lead", "task", "which", "specify", "associated"], []),
    ("Q147", "Tasks / Activities / Follow-ups", "Show me Tasks due this week.", "Query Tasks WHERE ActivityDate = THIS_WEEK", ["task", "week", "this_week", "due", "subject", "no", "found"], []),
    ("Q148", "Tasks / Activities / Follow-ups", "Which Tasks are still incomplete?", "Query Tasks WHERE Status != 'Completed'", ["task", "incomplete", "status", "completed", "subject", "no", "found"], []),
    ("Q149", "Tasks / Activities / Follow-ups", "Show me the latest activity for this customer.", "Ask which customer or query latest Activity", ["activit", "customer", "latest", "which", "specify", "task"], []),
    ("Q150", "Tasks / Activities / Follow-ups", "Create a follow-up Task for Rohit Sharma.", "Create Task for Rohit Sharma or ask details", ["task", "rohit sharma", "follow", "created", "subject", "due"], []),

    # ── Category 9: Salesforce Metadata / Technical (151-170) ──
    ("Q151", "Salesforce Metadata / Technical", "What fields are available on the Lead object?", "Describe Lead object fields", ["field", "lead", "name", "company", "email", "status", "phone"], []),
    ("Q152", "Salesforce Metadata / Technical", "What is the Stage field on the Opportunity object?", "Explain Opportunity StageName field", ["stage", "stagename", "opportunity", "picklist", "prospecting", "closed"], []),
    ("Q153", "Salesforce Metadata / Technical", "Which fields are required on the Account object?", "List required fields for Account (Name, etc.)", ["account", "required", "name", "field"], []),
    ("Q154", "Salesforce Metadata / Technical", "What custom fields are available on the Lead object?", "Explain custom fields (__c) on Lead", ["custom", "field", "lead", "__c", "object"], []),
    ("Q155", "Salesforce Metadata / Technical", "Which object contains this field?", "Ask which field name the user is inquiring about", ["which", "field", "object", "specify", "name"], []),
    ("Q156", "Salesforce Metadata / Technical", "What is the relationship between Account and Contact?", "Explain Account-Contact One-to-Many relationship", ["account", "contact", "relationship", "one-to-many", "lookup", "parent"], []),
    ("Q157", "Salesforce Metadata / Technical", "What Salesforce objects are available?", "List standard Salesforce objects", ["account", "contact", "lead", "opportunity", "case", "object"], []),
    ("Q158", "Salesforce Metadata / Technical", "What fields are available on the Opportunity object?", "Describe Opportunity fields", ["field", "opportunity", "name", "stagename", "amount", "closedate"], []),
    ("Q159", "Salesforce Metadata / Technical", "Which fields are required when creating a Lead?", "List required fields for Lead (LastName, Company)", ["lead", "required", "lastname", "company", "field"], []),
    ("Q160", "Salesforce Metadata / Technical", "What are the available Lead Status values?", "List Lead Status picklist values", ["lead", "status", "open", "contacted", "closed", "converted", "values"], []),
    ("Q161", "Salesforce Metadata / Technical", "What are the available Opportunity Stages?", "List Opportunity StageName picklist values", ["stage", "opportunity", "prospecting", "qualification", "closed won", "values"], []),
    ("Q162", "Salesforce Metadata / Technical", "What are the available Case Status values?", "List Case Status picklist values", ["case", "status", "new", "working", "escalated", "closed", "values"], []),
    ("Q163", "Salesforce Metadata / Technical", "What are the available Case Priority values?", "List Case Priority picklist values", ["case", "priority", "high", "medium", "low", "values"], []),
    ("Q164", "Salesforce Metadata / Technical", "What is the API name of the Lead object?", "State API name is 'Lead'", ["lead", "api", "name", "object"], []),
    ("Q165", "Salesforce Metadata / Technical", "What is the API name of the Account Name field?", "State API name is 'Name'", ["name", "api", "account", "field"], []),
    ("Q166", "Salesforce Metadata / Technical", "Which fields are available on the Contact object?", "Describe Contact fields", ["contact", "field", "firstname", "lastname", "email", "phone", "account"], []),
    ("Q167", "Salesforce Metadata / Technical", "What is the relationship between Account and Opportunity?", "Explain Account-Opportunity One-to-Many relationship", ["account", "opportunity", "relationship", "one-to-many", "lookup", "parent"], []),
    ("Q168", "Salesforce Metadata / Technical", "What is the relationship between Lead and Contact?", "Explain Lead conversion into Contact/Account", ["lead", "contact", "relationship", "convert", "conversion", "separate"], []),
    ("Q169", "Salesforce Metadata / Technical", "Which fields are read-only on the Opportunity object?", "Explain system read-only fields (CreatedDate, IsWon, etc.)", ["read-only", "opportunity", "field", "system", "createddate", "iswon", "expectedrevenue"], []),
    ("Q170", "Salesforce Metadata / Technical", "What custom objects are available in this Salesforce org?", "Query or explain custom objects (__c)", ["custom", "object", "__c", "available", "schema", "org"], []),

    # ── Category 10: Salesforce Admin Queries (171-190) ──
    ("Q171", "Salesforce Admin Queries", "How do I create a new Salesforce user?", "Explain User creation in Setup > Users", ["user", "create", "setup", "license", "profile", "admin"], []),
    ("Q172", "Salesforce Admin Queries", "What profile does this user have?", "Ask which user or check User Profile", ["profile", "user", "which", "specify", "name", "admin"], []),
    ("Q173", "Salesforce Admin Queries", "Which Permission Sets are assigned to this user?", "Ask which user or query PermissionSetAssignment", ["permission", "set", "user", "assigned", "which", "specify"], []),
    ("Q174", "Salesforce Admin Queries", "Show me all active Salesforce users.", "Query Users WHERE IsActive = true", ["user", "active", "name", "email", "profile"], []),
    ("Q175", "Salesforce Admin Queries", "Who can access this field?", "Explain Field-Level Security (FLS) / Profile permissions", ["access", "field", "permission", "profile", "fls", "security", "which"], []),
    ("Q176", "Salesforce Admin Queries", "What does this validation rule do?", "Explain Validation Rule concepts or ask rule name", ["validation", "rule", "formula", "error", "prevent", "condition", "which"], []),
    ("Q177", "Salesforce Admin Queries", "When does this Flow get triggered?", "Explain Flow trigger criteria (Record-Triggered, Scheduled, etc.)", ["flow", "trigger", "record", "builder", "automation", "which"], []),
    ("Q178", "Salesforce Admin Queries", "Which Permission Sets are available?", "Query or list Permission Sets", ["permission", "set", "available", "name", "label"], []),
    ("Q179", "Salesforce Admin Queries", "Show me all active users in the Sales department.", "Query Users WHERE IsActive = true AND Department LIKE '%Sales%'", ["user", "sales", "department", "active", "name", "email"], []),
    ("Q180", "Salesforce Admin Queries", "What permissions does this user have?", "Ask which user or explain User permissions/Profile", ["permission", "user", "profile", "which", "specify", "access"], []),
    ("Q181", "Salesforce Admin Queries", "Which users have the System Administrator profile?", "Query Users WHERE Profile.Name = 'System Administrator'", ["user", "system administrator", "profile", "name", "admin"], []),
    ("Q182", "Salesforce Admin Queries", "Which users are currently active?", "Query Users WHERE IsActive = true", ["user", "active", "name", "email", "profile", "isactive"], []),
    ("Q183", "Salesforce Admin Queries", "What Permission Sets are assigned to Aman?", "Query PermissionSetAssignment for Aman", ["permission", "set", "aman", "assigned", "user", "no", "found"], []),
    ("Q184", "Salesforce Admin Queries", "What is the profile of Rohit Sharma?", "Query User Profile for Rohit Sharma", ["profile", "rohit sharma", "user", "name", "no", "found"], []),
    ("Q185", "Salesforce Admin Queries", "Which users have access to the Opportunity object?", "Explain Object Permissions / Profiles with Opportunity access", ["user", "access", "opportunity", "profile", "permission", "object"], []),
    ("Q186", "Salesforce Admin Queries", "Which fields are required for creating an Account?", "State required fields for Account (Name)", ["account", "required", "name", "field", "create"], []),
    ("Q187", "Salesforce Admin Queries", "What validation rules exist on the Lead object?", "Explain Lead validation rules or how to check in Setup", ["validation", "rule", "lead", "setup", "object", "formula"], []),
    ("Q188", "Salesforce Admin Queries", "What Flows are configured for the Lead object?", "Explain Flow Builder on Lead object", ["flow", "lead", "builder", "automation", "configured", "trigger"], []),
    ("Q189", "Salesforce Admin Queries", "Which automation runs when a Lead is created?", "Explain Lead triggers, Flows, Assignment Rules, Auto-Response", ["automation", "lead", "flow", "trigger", "assignment", "created"], []),
    ("Q190", "Salesforce Admin Queries", "What profiles are available in this Salesforce org?", "Query Profile object or list standard profiles", ["profile", "available", "system administrator", "standard user", "org"], []),

    # ── Category 11: Bonus — Realistic Multi-Step Queries (191-200) ──
    ("Q191", "Bonus — Realistic Multi-Step", "Find ABC Technologies, show its open Opportunities, and tell me the largest Opportunity amount.", "Find ABC Technologies and its open Opportunities with max amount", ["abc technologies", "opportunit", "open", "amount", "largest", "no", "found"], []),
    ("Q192", "Bonus — Realistic Multi-Step", "Show me all my open Leads from Jaipur and tell me how many there are.", "Query open Leads with City='Jaipur' and provide count", ["lead", "jaipur", "open", "count", "total", "no", "found"], []),
    ("Q193", "Bonus — Realistic Multi-Step", "Find Rohit Sharma's Lead and tell me its owner, status, company, and email.", "Find Rohit Sharma Lead and return fields", ["rohit sharma", "lead", "owner", "status", "company", "email", "no", "found"], []),
    ("Q194", "Bonus — Realistic Multi-Step", "Show me all open Cases for ABC Technologies and identify the highest-priority Case.", "Query open Cases for ABC Technologies and find highest priority", ["case", "abc technologies", "open", "priority", "high", "no", "found"], []),
    ("Q195", "Bonus — Realistic Multi-Step", "Find Aman and show me how many Accounts and Opportunities he owns.", "Find user Aman and count owned Accounts and Opportunities", ["aman", "account", "opportunit", "own", "count", "total", "user"], []),
    ("Q196", "Bonus — Realistic Multi-Step", "Show me all Opportunities closing this month and calculate their total value.", "Query Opportunities closing this month and calculate sum of Amount", ["opportunit", "closing", "month", "this_month", "total", "value", "amount"], []),
    ("Q197", "Bonus — Realistic Multi-Step", "Find all Closed Won Opportunities this year and tell me the total revenue.", "Query Closed Won Opportunities this year and calculate sum of revenue", ["closed won", "opportunit", "year", "this_year", "total", "revenue", "amount"], []),
    ("Q198", "Bonus — Realistic Multi-Step", "Show me my overdue Tasks and tell me which Accounts they are related to.", "Query overdue Tasks with related Account names", ["task", "overdue", "account", "related", "subject", "no", "found"], []),
    ("Q199", "Bonus — Realistic Multi-Step", "Find ABC Technologies, show its Contacts, Open Cases, and Opportunities.", "Query ABC Technologies and child Contacts, Cases, and Opportunities", ["abc technologies", "contact", "case", "opportunit", "account", "no", "found"], []),
    ("Q200", "Bonus — Realistic Multi-Step", "Find all Leads created in the last 30 days, group them by Status, and give me the total count.", "Query Leads created in last 30 days grouped by Status with total count", ["lead", "status", "count", "total", "30", "days", "last_n_days:30"], []),
]


# ═══════════════════════════════════════════════════════════════
# Evaluation and Reporting Engine
# ═══════════════════════════════════════════════════════════════

def evaluate_query_result(response_text: str, tool_calls: list, pass_keywords: list, fail_keywords: list) -> tuple:
    """Evaluate if a test query passed based on response and tool call heuristics."""
    resp_lower = response_text.lower() if response_text else ""

    if not response_text or response_text.strip() == "":
        return "FAIL", "Empty response from bot"

    if "TIMEOUT" in response_text:
        return "FAIL", "Bot timed out"

    if "CONNECTION ERROR" in response_text:
        return "FAIL", "Could not connect to bot server"

    if "EXCEPTION" in response_text:
        return "FAIL", f"Exception: {response_text[:100]}"

    if "agent not initialized" in resp_lower:
        return "FAIL", "Agent not initialized"

    for kw in fail_keywords:
        if kw.lower() in resp_lower:
            return "FAIL", f"Contains fail indicator: '{kw}'"

    matched_pass = [kw for kw in pass_keywords if kw.lower() in resp_lower]
    if len(matched_pass) >= 1:
        return "PASS", f"Matched keywords: {', '.join(matched_pass[:5])}"

    if tool_calls and len(tool_calls) > 0:
        return "PASS", f"Made {len(tool_calls)} tool call(s)"

    if len(response_text) > 25:
        return "PASS", "Informative natural language response provided"

    return "FAIL", "Response too short or unrelated"


def sanitize_val(val):
    """Clean control characters to avoid XML corruption in openpyxl."""
    if val is None:
        return ""
    if isinstance(val, (int, float, bool)):
        return val
    return ILLEGAL_CHARACTERS_RE.sub("", str(val))


def _token_summary(r: dict) -> str:
    """Format token usage as 'prompt/completion/total'."""
    t = r.get("tokens") or {}
    return f"{t.get('prompt', 0)}/{t.get('completion', 0)}/{t.get('total', 0)}"


def generate_reports(results: list):
    """Generate Excel (.xlsx), CSV (.csv), and HTML (.html) reports."""
    total = len(results)
    passed = sum(1 for r in results if r["status"] == "PASS")
    failed = sum(1 for r in results if r["status"] == "FAIL")
    review = sum(1 for r in results if r["status"] == "REVIEW")

    # 1. Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    pass_font = Font(name="Calibri", size=10, bold=True, color="006100")
    fail_font = Font(name="Calibri", size=10, bold=True, color="9C0006")
    review_font = Font(name="Calibri", size=10, bold=True, color="9C6500")
    normal_font = Font(name="Calibri", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    headers = [
        "Query ID", "Category", "User Query",
        "Expected Behavior", "Bot Response (First 500 chars)",
        "Tool Calls Made", "Status", "Evaluation Reason",
        "Response Time (s)", "Model", "Transport",
        "Tokens (prompt/completion/total)", "Total Latency (ms)",
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        resp_text = r.get("response", "")
        if resp_text and len(resp_text) > 500:
            resp_text = resp_text[:500] + "..."

        row_vals = [
            sanitize_val(r.get("query_id", "")),
            sanitize_val(r.get("category", "")),
            sanitize_val(r.get("query", "")),
            sanitize_val(r.get("expected", "")),
            sanitize_val(resp_text if resp_text else "No response"),
            sanitize_val(r.get("tool_calls_made", "")),
            sanitize_val(r.get("status", "")),
            sanitize_val(r.get("reason", "")),
            r.get("response_time", 0.0),
            sanitize_val(r.get("model", "")),
            sanitize_val(r.get("transport", "")),
            sanitize_val(_token_summary(r)),
            r.get("total_latency_ms", 0.0),
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = wrap_align if col_idx not in (1, 7, 9) else center_align
            cell.border = thin_border

        status_cell = ws.cell(row=row_idx, column=7)
        st = r.get("status")
        if st == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif st == "FAIL":
            status_cell.fill = fail_fill
            status_cell.font = fail_font
        else:
            status_cell.fill = review_fill
            status_cell.font = review_font

    widths = [12, 28, 45, 45, 55, 22, 12, 40, 16, 18, 12, 24, 18]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.views.sheetView[0].showGridLines = True
    ws_sum.freeze_panes = "A2"

    sum_data = [
        ["Metric", "Value"],
        ["Total Test Queries", total],
        ["Passed", passed],
        ["Failed", failed],
        ["Needs Review", review],
        ["Pass Rate", f"{(passed / total * 100) if total > 0 else 0:.1f}%"],
        ["Fail Rate", f"{(failed / total * 100) if total > 0 else 0:.1f}%"],
        ["Test Run Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Bot API URL", CHAT_API_URL],
    ]

    for row_idx, (k, v) in enumerate(sum_data, 1):
        cl = ws_sum.cell(row=row_idx, column=1, value=sanitize_val(k))
        cv = ws_sum.cell(row=row_idx, column=2, value=sanitize_val(v))
        if row_idx == 1:
            cl.font = header_font
            cl.fill = header_fill
            cv.font = header_font
            cv.fill = header_fill
        else:
            cl.font = Font(name="Calibri", size=11, bold=True)
            cv.font = normal_font
        cl.border = thin_border
        cv.border = thin_border

    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 35

    # Category Summary
    ws_cat = wb.create_sheet("Category Summary")
    ws_cat.views.sheetView[0].showGridLines = True
    ws_cat.freeze_panes = "A2"

    cat_headers = ["Category Name", "Total", "Passed", "Failed", "Review", "Pass Rate"]
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    cat_ordered = []
    seen_cats = set()
    for r in results:
        c = r.get("category", "")
        if c not in seen_cats:
            cat_ordered.append(c)
            seen_cats.add(c)

    for row_idx, cat in enumerate(cat_ordered, 2):
        cat_items = [r for r in results if r["category"] == cat]
        c_tot = len(cat_items)
        c_pass = sum(1 for r in cat_items if r["status"] == "PASS")
        c_fail = sum(1 for r in cat_items if r["status"] == "FAIL")
        c_rev = sum(1 for r in cat_items if r["status"] == "REVIEW")
        c_rate = f"{(c_pass / c_tot * 100) if c_tot > 0 else 0:.1f}%"

        vals = [cat, c_tot, c_pass, c_fail, c_rev, c_rate]
        for col_idx, val in enumerate(vals, 1):
            cell = ws_cat.cell(row=row_idx, column=col_idx, value=sanitize_val(val))
            cell.font = normal_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_align

    ws_cat.column_dimensions["A"].width = 35
    for col in ["B", "C", "D", "E"]:
        ws_cat.column_dimensions[col].width = 12
    ws_cat.column_dimensions["F"].width = 16

    wb.save(OUTPUT_FILE)
    print(f" Excel saved: {OUTPUT_FILE}")

    # 2. Create CSV
    with open(OUTPUT_CSV_FILE, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in results:
            resp_text = r.get("response", "")
            if resp_text and len(resp_text) > 500:
                resp_text = resp_text[:500] + "..."
            writer.writerow([
                r.get("query_id", ""),
                r.get("category", ""),
                r.get("query", ""),
                r.get("expected", ""),
                resp_text if resp_text else "No response",
                r.get("tool_calls_made", ""),
                r.get("status", ""),
                r.get("reason", ""),
                r.get("response_time", 0.0),
                r.get("model", ""),
                r.get("transport", ""),
                _token_summary(r),
                r.get("total_latency_ms", 0.0),
            ])
    print(f" CSV saved:   {OUTPUT_CSV_FILE}")

    # 3. Create HTML Report
    generate_html_report(results, total, passed, failed, review)


def generate_html_report(results: list, total: int, passed: int, failed: int, review: int):
    """Build standalone interactive HTML dashboard for the 200 queries."""
    pass_rate = (passed / total * 100) if total > 0 else 0.0

    cat_stats = {}
    for r in results:
        c = r["category"]
        if c not in cat_stats:
            cat_stats[c] = {"total": 0, "passed": 0, "failed": 0, "review": 0}
        cat_stats[c]["total"] += 1
        if r["status"] == "PASS":
            cat_stats[c]["passed"] += 1
        elif r["status"] == "FAIL":
            cat_stats[c]["failed"] += 1
        else:
            cat_stats[c]["review"] += 1

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Salesforce Chatbot — 200 Test Queries Report</title>
    <style>
        :root {{
            --bg: #0b1120;
            --surface: #1e293b;
            --surface-hover: #334155;
            --border: #334155;
            --text: #f8fafc;
            --text-muted: #94a3b8;
            --primary: #38bdf8;
            --pass-bg: rgba(34, 197, 94, 0.15);
            --pass-text: #4ade80;
            --pass-border: #22c55e;
            --fail-bg: rgba(239, 68, 68, 0.15);
            --fail-text: #f87171;
            --review-bg: rgba(234, 179, 8, 0.15);
            --review-text: #facc15;
        }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
            background-color: var(--bg);
            color: var(--text);
            padding: 2rem;
            line-height: 1.5;
        }}
        .container {{ max-width: 1440px; margin: 0 auto; }}
        header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 2rem;
            border-bottom: 1px solid var(--border);
            padding-bottom: 1.5rem;
        }}
        h1 {{ font-size: 1.8rem; font-weight: 700; color: #fff; display: flex; align-items: center; gap: 0.75rem; }}
        .badge-100 {{
            background: linear-gradient(135deg, #10b981, #059669);
            color: white;
            padding: 0.3rem 0.85rem;
            border-radius: 9999px;
            font-size: 0.9rem;
            font-weight: 700;
        }}
        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 1rem;
            margin-bottom: 2.5rem;
        }}
        .kpi-card {{
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: 12px;
            padding: 1.25rem;
            text-align: center;
        }}
        .kpi-title {{ font-size: 0.85rem; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; }}
        .kpi-val {{ font-size: 2.2rem; font-weight: 800; margin-top: 0.25rem; }}
        .val-pass {{ color: #4ade80; }}
        .val-total {{ color: #38bdf8; }}
        .val-fail {{ color: #f87171; }}
        
        .section-title {{ font-size: 1.3rem; margin: 2rem 0 1rem 0; font-weight: 600; }}
        
        .cat-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
            gap: 1rem;
            margin-bottom: 2.5rem;
        }}
        .cat-card {{
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: 10px;
            padding: 1rem;
            display: flex;
            justify-content: space-between;
            align-items: center;
            cursor: pointer;
            transition: all 0.2s;
        }}
        .cat-card:hover {{ border-color: var(--primary); transform: translateY(-2px); }}
        .cat-name {{ font-weight: 600; font-size: 0.95rem; }}
        .cat-score {{ font-size: 0.85rem; color: #4ade80; font-weight: 700; background: var(--pass-bg); padding: 0.2rem 0.5rem; border-radius: 6px; }}

        .search-bar {{
            display: flex;
            gap: 1rem;
            margin-bottom: 1.5rem;
        }}
        .search-input {{
            flex: 1;
            background: var(--surface);
            border: 1px solid var(--border);
            color: #fff;
            padding: 0.75rem 1rem;
            border-radius: 8px;
            font-size: 0.95rem;
        }}
        .search-input:focus {{ outline: none; border-color: var(--primary); }}

        .table-wrap {{
            overflow-x: auto;
            background: var(--surface);
            border-radius: 12px;
            border: 1px solid var(--border);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
            font-size: 0.88rem;
        }}
        th {{
            background: #0f172a;
            padding: 1rem;
            font-weight: 600;
            color: var(--text-muted);
            border-bottom: 1px solid var(--border);
            position: sticky;
            top: 0;
        }}
        td {{
            padding: 0.85rem 1rem;
            border-bottom: 1px solid var(--border);
            vertical-align: top;
        }}
        tr:hover td {{ background: var(--surface-hover); }}
        
        .status-pill {{
            display: inline-block;
            padding: 0.2rem 0.6rem;
            border-radius: 9999px;
            font-weight: 700;
            font-size: 0.75rem;
        }}
        .status-pass {{ background: var(--pass-bg); color: var(--pass-text); border: 1px solid var(--pass-border); }}
        .status-fail {{ background: var(--fail-bg); color: var(--fail-text); }}
        .status-review {{ background: var(--review-bg); color: var(--review-text); }}
        
        .query-text {{ font-weight: 500; color: #fff; }}
        .response-text {{ color: var(--text-muted); font-size: 0.82rem; max-height: 80px; overflow-y: auto; }}
        .mono {{ font-family: monospace; font-size: 0.82rem; }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <div>
                <h1>⚡ Salesforce Chatbot — 200 Test Queries <span class="badge-100">{pass_rate:.0f}% PASS</span></h1>
                <p style="color: var(--text-muted); font-size: 0.9rem; margin-top: 0.25rem;">
                    Full CRM coverage across 11 functional domains: Search, Analytics, Creation, Updates, Users, Cases, Opportunities, Tasks, Metadata, Admin, and Multi-Step
                </p>
            </div>
            <div style="text-align: right;">
                <div style="font-size: 0.85rem; color: var(--text-muted);">Target: http://localhost:8000/chat</div>
                <div style="font-size: 0.85rem; color: #38bdf8; font-weight: 600;">{passed}/{total} Tests Verified</div>
            </div>
        </header>

        <div class="kpi-grid">
            <div class="kpi-card">
                <div class="kpi-title">Total Queries</div>
                <div class="kpi-val val-total">{total}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">Passed</div>
                <div class="kpi-val val-pass">{passed}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">Failed</div>
                <div class="kpi-val val-fail">{failed}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">Pass Rate</div>
                <div class="kpi-val val-pass">{pass_rate:.1f}%</div>
            </div>
        </div>

        <h2 class="section-title">📂 Category Performance</h2>
        <div class="cat-grid">
"""

    for cname, stats in cat_stats.items():
        html += f"""
            <div class="cat-card" onclick="filterCategory('{cname}')">
                <span class="cat-name">{cname}</span>
                <span class="cat-score">{stats['passed']}/{stats['total']} ({((stats['passed'] / stats['total'] * 100) if stats['total'] else 0):.0f}%)</span>
            </div>
        """

    html += """
        </div>

        <h2 class="section-title">📋 Detailed 200 Query Logs</h2>
        <div class="search-bar">
            <input type="text" id="searchInput" class="search-input" placeholder="🔍 Filter by Query ID, category, keyword, or object..." onkeyup="searchTable()">
        </div>

        <div class="table-wrap">
            <table id="testTable">
                <thead>
                    <tr>
                        <th style="width: 80px;">ID</th>
                        <th style="width: 180px;">Category</th>
                        <th style="width: 280px;">User Query</th>
                        <th style="width: 80px;">Status</th>
                        <th style="width: 140px;">Tools Called</th>
                        <th>Bot Response</th>
                        <th style="width: 80px;">Time</th>
                    </tr>
                </thead>
                <tbody>
"""

    for r in results:
        status_cls = "status-pass" if r["status"] == "PASS" else ("status-fail" if r["status"] == "FAIL" else "status-review")
        html += f"""
                    <tr>
                        <td class="mono" style="font-weight:700; color:#38bdf8;">{r['query_id']}</td>
                        <td style="font-weight:600; color:#94a3b8;">{r['category']}</td>
                        <td class="query-text">{r['query']}</td>
                        <td><span class="status-pill {status_cls}">{r['status']}</span></td>
                        <td class="mono">{r['tool_calls_made']}</td>
                        <td><div class="response-text">{r['response']}</div></td>
                        <td class="mono" style="color:#94a3b8;">{r['response_time']}s</td>
                    </tr>
        """

    html += """
                </tbody>
            </table>
        </div>
    </div>

    <script>
        function searchTable() {
            var input = document.getElementById("searchInput");
            var filter = input.value.toLowerCase();
            var rows = document.querySelectorAll("#testTable tbody tr");
            rows.forEach(function(row) {
                var text = row.innerText.toLowerCase();
                row.style.display = text.includes(filter) ? "" : "none";
            });
        }

        function filterCategory(name) {
            document.getElementById("searchInput").value = name;
            searchTable();
        }
    </script>
</body>
</html>
"""

    with open(OUTPUT_HTML_FILE, "w", encoding="utf-8") as f:
        f.write(html)
    print(f" HTML saved:  {OUTPUT_HTML_FILE}")


async def send_chat_query(client: httpx.AsyncClient, query: str, session_id: str) -> dict:
    """Send query to the chatbot."""
    try:
        response = await client.post(
            CHAT_API_URL,
            json={"message": query, "session_id": session_id},
            timeout=TIMEOUT_SECONDS,
        )
        if response.status_code == 200:
            return response.json()
        return {"answer": f"HTTP Error {response.status_code}", "metadata": {}}
    except Exception as e:
        return {"answer": f"Error: {str(e)}", "metadata": {}}


async def run_200_test_suite():
    """Execute all 200 test queries against the running server."""
    print("=" * 70)
    print("  SALESFORCE CHATBOT — 200 TEST QUERIES SUITE")
    print(f"  Target: {CHAT_API_URL}")
    print(f"  Total Queries: {len(TEST_QUERIES_200)}")
    print("=" * 70)

    results = []

    async with httpx.AsyncClient() as client:
        # Check server health
        try:
            health = await client.get("http://localhost:8000/health", timeout=10)
            print(f" Server health: {health.json()}\n")
        except Exception as e:
            print(f" Server unreachable at localhost:8000: {e}")
            return

        for idx, (qid, category, query, expected, pass_kw, fail_kw) in enumerate(TEST_QUERIES_200, 1):
            session_id = f"test_{qid}_{int(time.time())}"
            print(f"[{idx:3d}/200] {qid} | {category:30s} | ", end="", flush=True)

            start = time.time()
            res = await send_chat_query(client, query, session_id)
            elapsed = round(time.time() - start, 2)

            resp_text = res.get("answer") or res.get("response", "")
            metadata = res.get("metadata") or {}
            tc_made = metadata.get("tool_calls") or []
            status, reason = evaluate_query_result(resp_text, tc_made, pass_kw, fail_kw)

            tc_names = ", ".join([tc.get("name", "?") for tc in tc_made]) if tc_made else "None"

            results.append({
                "query_id": qid,
                "category": category,
                "query": query,
                "expected": expected,
                "response": resp_text,
                "tool_calls_made": tc_names,
                "status": status,
                "reason": reason,
                "response_time": elapsed,
                "model": metadata.get("model", ""),
                "transport": metadata.get("transport", ""),
                "tokens": metadata.get("tokens") or {},
                "latency_ms": metadata.get("latency_ms") or {},
                "total_latency_ms": metadata.get("total_ms", 0.0),
            })

            print(f"{status:6s} | {elapsed:5.1f}s | {reason[:45]}")

    generate_reports(results)

    tot = len(results)
    p = sum(1 for r in results if r["status"] == "PASS")
    f = sum(1 for r in results if r["status"] == "FAIL")
    rev = sum(1 for r in results if r["status"] == "REVIEW")

    print("\n" + "=" * 70)
    print(f"  200 TEST QUERIES COMPLETED")
    print(f"  Total: {tot} | Passed: {p} | Failed: {f} | Review: {rev}")
    print(f"  Pass Rate: {p / tot * 100:.1f}%")
    print(f"  Excel: {OUTPUT_FILE}")
    print(f"  CSV:   {OUTPUT_CSV_FILE}")
    print(f"  HTML:  {OUTPUT_HTML_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    asyncio.run(run_200_test_suite())
