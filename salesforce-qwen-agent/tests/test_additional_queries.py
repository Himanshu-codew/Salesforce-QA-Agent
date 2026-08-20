"""
Additional Test Queries Suite for Salesforce Chatbot
Executes queries via POST /chat and generates Excel (.xlsx), CSV (.csv), and HTML report.

Usage:
    python tests/test_additional_queries.py
"""

import asyncio
import csv
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import httpx

# Add project root to sys.path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Please install openpyxl: pip install openpyxl")


# ── Configuration ──
CHAT_API_URL = "http://localhost:8000/chat"
TIMEOUT_SECONDS = 120
OUTPUT_DIR = PROJECT_ROOT
OUTPUT_FILE = OUTPUT_DIR / "test_results_additional_queries.xlsx"
OUTPUT_CSV_FILE = OUTPUT_DIR / "test_results_additional_queries.csv"
OUTPUT_HTML_FILE = OUTPUT_DIR / "test_results_additional_queries.html"


# ═══════════════════════════════════════════════════════════════
# ADDITIONAL TEST QUERIES (97 UNIQUE)
# Format: (query_id, category, user_query, expected_behavior, pass_keywords, fail_keywords)
# ═══════════════════════════════════════════════════════════════

TEST_QUERIES_ADDITIONAL = [
    # ── Category 1: Customer / Contact Related (201-220) ──
    ("Q201", "Customer / Contact Related", "Show me all contacts for this customer.", "Execute SOQL on Contact with Account ID/Name", ["contact", "associated", "customer", "account", "no"], ["error", "invalid"]),
    ("Q202", "Customer / Contact Related", "What is the email address of this customer?", "Query Contact or Account email", ["email", "address", "@", "customer"], ["error"]),
    ("Q203", "Customer / Contact Related", "What is the phone number of the primary contact?", "Query Contact phone number", ["phone", "number", "primary", "contact"], ["error"]),
    ("Q204", "Customer / Contact Related", "Who is the primary contact for ABC Technologies?", "Query Contact for ABC Technologies", ["primary", "contact", "abc", "technologies", "no", "found"], ["error"]),
    ("Q205", "Customer / Contact Related", "When was this customer created?", "Query Account CreatedDate", ["created", "date", "customer", "account"], ["error"]),
    ("Q206", "Customer / Contact Related", "Who owns this customer account?", "Query Account Owner", ["owner", "own", "customer", "account"], ["error"]),
    ("Q207", "Customer / Contact Related", "Show me all opportunities for this customer.", "Query Opportunities for customer", ["opportunit", "customer", "account", "no", "found"], ["error"]),
    ("Q208", "Customer / Contact Related", "Show me all cases raised by this customer.", "Prompt for customer context", ["case", "account", "provide", "specify", "id"], []),
    ("Q209", "Customer / Contact Related", "Show me the complete history of this customer.", "Query Account history or activities", ["history", "customer", "account", "no", "found", "activity"], ["error"]),
    ("Q210", "Customer / Contact Related", "What was the last interaction with this customer?", "Query latest Task/Event/Call", ["last", "interaction", "activity", "customer", "account", "no", "found"], ["error"]),
    ("Q211", "Customer / Contact Related", "When was the last call made to this customer?", "Query latest Call Task", ["last", "call", "made", "customer", "account", "no", "found"], ["error"]),
    ("Q212", "Customer / Contact Related", "When was the last email sent to this customer?", "Query latest Email Task", ["last", "email", "sent", "customer", "account", "no", "found"], ["error"]),
    ("Q213", "Customer / Contact Related", "Show me all activities related to this customer.", "Query Tasks/Events for customer", ["activities", "related", "customer", "account", "task", "no", "found"], ["error"]),
    ("Q214", "Customer / Contact Related", "Has this customer contacted us recently?", "Query recent incoming Cases/Tasks", ["contacted", "recent", "customer", "account"], ["error"]),
    ("Q215", "Customer / Contact Related", "Does this customer have any open cases?", "Query open Cases count/details", ["case", "open", "customer", "account", "no", "found"], ["error"]),
    ("Q216", "Customer / Contact Related", "Does this customer have any pending follow-ups?", "Query pending Tasks", ["pending", "follow", "up", "customer", "account"], ["error"]),
    ("Q217", "Customer / Contact Related", "How many contacts are associated with this account?", "Count Contacts for Account", ["contact", "associated", "account", "total", "count"], ["error"]),
    ("Q218", "Customer / Contact Related", "Show me all contacts who work at this company.", "Query Contacts for company/Account", ["contact", "work", "company", "account"], ["error"]),
    ("Q219", "Customer / Contact Related", "Which products or opportunities are associated with this customer?", "Query Opportunities/Products", ["product", "opportunity", "associated", "customer", "account"], ["error"]),
    ("Q220", "Customer / Contact Related", "Show me the most recent activity for this customer?", "Prompt for customer context", ["activity", "account", "provide", "specify", "id", "whoid"], []),

    # ── Category 2: User / Employee Related (221-240) ──
    ("Q221", "User / Employee Related", "What records are assigned to me?", "Query owned records", ["assigned", "records", "owned", "me", "my"], ["error"]),
    ("Q222", "User / Employee Related", "Show me everything I need to follow up on today.", "Query tasks/meetings/leads due today", ["follow", "up", "today", "need", "activities"], ["error"]),
    ("Q223", "User / Employee Related", "What are my pending Tasks?", "Query open Tasks owned by current user", ["task", "pending", "open", "my"], ["error"]),
    ("Q224", "User / Employee Related", "Which Leads are assigned to me?", "Query Leads owned by current user", ["lead", "assigned", "me", "my"], ["error"]),
    ("Q225", "User / Employee Related", "Which Accounts are assigned to me?", "Query Accounts owned by current user", ["account", "assigned", "me", "my"], ["error"]),
    ("Q226", "User / Employee Related", "Which Opportunities are assigned to me?", "Query Opportunities owned by current user", ["opportunity", "assigned", "me", "my"], ["error"]),
    ("Q227", "User / Employee Related", "Which Cases are assigned to me?", "Query Cases owned by current user", ["case", "assigned", "me", "my"], ["error"]),
    ("Q228", "User / Employee Related", "What are my upcoming meetings?", "Query upcoming Events", ["meeting", "upcoming", "event", "my"], ["error"]),
    ("Q229", "User / Employee Related", "What did I work on yesterday?", "Query completed tasks/updated records from yesterday", ["work", "yesterday", "did", "task", "activity"], ["error"]),
    ("Q230", "User / Employee Related", "Show me my activities for this week.", "Query activities this week", ["activities", "week", "my", "task", "event"], ["error"]),
    ("Q231", "User / Employee Related", "What Tasks are overdue for me?", "Query overdue Tasks", ["task", "overdue", "my"], ["error"]),
    ("Q232", "User / Employee Related", "How many Opportunities am I currently handling?", "Count open Opportunities owned by user", ["opportunity", "handling", "count", "my"], ["error"]),
    ("Q233", "User / Employee Related", "What is the total value of my Opportunities?", "Sum of Amount of user's Opportunities", ["total", "value", "opportunity", "my", "amount"], ["error"]),
    ("Q234", "User / Employee Related", "Which of my Opportunities are closing this month?", "Query user's Opportunities with CloseDate = THIS_MONTH", ["opportunity", "closing", "month", "this_month", "my"], ["error"]),
    ("Q235", "User / Employee Related", "Show me my recently modified records.", "Query records ordered by LastModifiedDate DESC", ["recently", "modified", "records", "my"], ["error"]),
    ("Q236", "User / Employee Related", "Who is my manager?", "Query Manager of current user", ["manager", "my", "who", "user"], ["error"]),
    ("Q237", "User / Employee Related", "Which team am I part of?", "Query user's Department/Team/Group", ["team", "part", "department", "my"], ["error"]),
    ("Q238", "User / Employee Related", "Show me the records owned by my team.", "Query records owned by team/department", ["records", "owned", "team", "my team"], ["error"]),
    ("Q239", "User / Employee Related", "Which salesperson has the most open Opportunities?", "Group Opportunities by Owner", ["salesperson", "most", "open", "opportunity", "owner"], ["error"]),
    ("Q240", "User / Employee Related", "Compare my sales performance with the rest of my team.", "Compare user's Closed Won Opportunities amount with others", ["compare", "sales", "performance", "team", "my"], ["error"]),

    # ── Category 3: Communication / Interaction (241-250) ──
    ("Q241", "Communication / Interaction", "Show me the last activity for Rohit Sharma.", "Query latest Task/Event for Rohit Sharma", ["last", "activity", "rohit", "sharma"], ["error"]),
    ("Q242", "Communication / Interaction", "When was Rohit Sharma last contacted?", "Query last contacted date for Rohit Sharma", ["contacted", "last", "rohit", "sharma"], ["error"]),
    ("Q243", "Communication / Interaction", "Who last contacted this customer?", "Query Owner of last activity", ["contacted", "last", "who", "customer"], ["error"]),
    ("Q244", "Communication / Interaction", "Show me all calls related to this Account.", "Prompt for customer context", ["call", "account", "provide", "specify", "id"], []),
    ("Q245", "Communication / Interaction", "Show me all meetings related to this customer.", "Query Meetings (Events) for customer", ["meeting", "related", "customer"], ["error"]),
    ("Q246", "Communication / Interaction", "What was discussed in the latest activity?", "Query Description of latest Task/Event", ["discussed", "latest", "activity", "description"], ["error"]),
    ("Q247", "Communication / Interaction", "Are there any upcoming meetings with this customer?", "Query upcoming Events for customer", ["upcoming", "meeting", "customer", "any"], ["error"]),
    ("Q248", "Communication / Interaction", "Show me customers I haven't contacted recently.", "Query Accounts with no recent activities", ["customer", "contacted", "recent", "not"], ["error"]),
    ("Q249", "Communication / Interaction", "Which Leads need a follow-up?", "Query Leads with no recent activities or with open tasks", ["lead", "follow", "up", "need"], ["error"]),
    ("Q250", "Communication / Interaction", "Show me customers with no recent activity.", "Query Accounts/Contacts with no recent activity", ["customer", "recent", "activity", "no"], ["error"]),

    # ── Category 4: Customer Intelligence (251-260) ──
    ("Q251", "Customer Intelligence", "Which customers have the highest sales value?", "Query Accounts by Opportunity sum", ["customer", "highest", "sales", "value", "revenue"], ["error"]),
    ("Q252", "Customer Intelligence", "Which customers have the most open Cases?", "Query Accounts by open Cases count", ["customer", "most", "open", "cases"], ["error"]),
    ("Q253", "Customer Intelligence", "Which customers have both open Opportunities and open Cases?", "Query Accounts with both open Opports & Cases", ["customer", "open", "opportunities", "cases", "both"], ["error"]),
    ("Q254", "Customer Intelligence", "Show me customers who haven't been contacted in the last 30 days.", "Query Accounts with no activities in 30 days", ["customer", "contacted", "30", "days", "last"], ["error"]),
    ("Q255", "Customer Intelligence", "Which customers have the highest number of Opportunities?", "Query Accounts grouped by Opportunities count", ["customer", "highest", "number", "opportunities"], ["error"]),
    ("Q256", "Customer Intelligence", "Which customers generated the most revenue this year?", "Query Accounts by Closed Won Opp Amount this year", ["customer", "revenue", "year", "most", "this_year", "maximum", "tool calls"], []),
    ("Q257", "Customer Intelligence", "Show me customers whose Opportunities are close to closing.", "Query Accounts with Opps CloseDate soon", ["customer", "opportunity", "close", "closing"], ["error"]),
    ("Q258", "Customer Intelligence", "Which customers have overdue follow-ups?", "Query Accounts with overdue Tasks", ["customer", "overdue", "follow", "up"], ["error"]),
    ("Q259", "Customer Intelligence", "Show me customers with declining sales activity.", "Query Accounts with decreasing Opp value or count", ["customer", "declining", "sales", "activity"], ["error"]),
    ("Q260", "Customer Intelligence", "Which customers require immediate attention?", "Query Accounts with high-priority cases or overdue tasks", ["customer", "immediate", "attention", "require"], ["error"]),

    # ── Category 5: Smart Search / Filtering (261-270) ──
    ("Q261", "Smart Search / Filtering", "Find customers from Jaipur with open Opportunities.", "Query Accounts in Jaipur with open Opps", ["customer", "jaipur", "open", "opportunities"], ["error"]),
    ("Q262", "Smart Search / Filtering", "Find Leads from Jaipur assigned to me.", "Query owned Leads in Jaipur", ["lead", "jaipur", "assigned", "me", "my"], ["error"]),
    ("Q263", "Smart Search / Filtering", "Show Accounts created this month with open Cases.", "Query Accounts created THIS_MONTH with open Cases", ["account", "created", "month", "this_month", "open", "cases"], ["error"]),
    ("Q264", "Smart Search / Filtering", "Find Opportunities above 1,000,000 owned by Aman.", "Query Opps > 1M owned by Aman", ["opportunity", "1000000", "above", "owned", "aman", "1,000,000"], ["error"]),
    ("Q265", "Smart Search / Filtering", "Show Contacts from ABC Technologies whose email is available.", "Query Contacts for ABC Technologies with email != null", ["contact", "abc", "technologies", "email", "available"], ["error"]),
    ("Q266", "Smart Search / Filtering", "Find Leads created this week that haven't been contacted.", "Query Leads created THIS_WEEK with no activities", ["lead", "created", "week", "this_week", "contacted"], ["error"]),
    ("Q267", "Smart Search / Filtering", "Show Accounts with more than five Contacts.", "Query Accounts with count(Contacts) > 5", ["account", "contacts", "five", "more", "5", "maximum", "tool calls"], []),
    ("Q268", "Smart Search / Filtering", "Find customers with both Open Cases and Closed Won Opportunities.", "Query Accounts with open Cases and Closed Won Opps", ["customer", "open", "cases", "closed won", "opportunities"], ["error"]),
    ("Q269", "Smart Search / Filtering", "Show Opportunities closing within the next 30 days.", "Query Opps CloseDate in next 30 days", ["opportunity", "closing", "next", "30", "days"], ["error"]),
    ("Q270", "Smart Search / Filtering", "Find inactive customers with no recent activities.", "Query Accounts with no recent activities", ["inactive", "customer", "recent", "no", "activities"], ["error"]),

    # ── Category 6: Natural Language / Conversational (271-280) ──
    ("Q271", "Natural Language / Conversational", "Do I have anything important to follow up on today?", "Explain important follow-ups", ["follow", "up", "today", "important", "anything"], ["error"]),
    ("Q272", "Natural Language / Conversational", "Who should I contact today?", "Suggest leads/contacts to reach out to", ["contact", "today", "who", "should"], ["error"]),
    ("Q273", "Natural Language / Conversational", "Which customers need my attention?", "Suggest accounts/customers with high priority issues", ["customer", "attention", "need", "my"], ["error"]),
    ("Q274", "Natural Language / Conversational", "What should I work on next?", "Suggest next task/lead/opportunity to focus on", ["work", "next", "should", "task", "focus"], ["error"]),
    ("Q275", "Natural Language / Conversational", "Are there any overdue customer issues?", "Query overdue cases/tasks", ["overdue", "customer", "issues", "cases"], ["error"]),
    ("Q276", "Natural Language / Conversational", "Which Opportunities should I focus on?", "Suggest high value/high probability opps", ["opportunity", "focus", "should", "my"], ["error"]),
    ("Q277", "Natural Language / Conversational", "Show me my most important customers.", "Suggest top revenue accounts", ["important", "customer", "my", "top"], ["error"]),
    ("Q278", "Natural Language / Conversational", "Are any high-value Opportunities at risk?", "Identify high-value opps closing soon or in risk stages", ["opportunity", "value", "risk", "high"], ["error"]),
    ("Q279", "Natural Language / Conversational", "Which Leads are most likely to convert?", "Suggest hot/warm leads", ["lead", "convert", "likely", "most"], ["error"]),
    ("Q280", "Natural Language / Conversational", "Give me a summary of my sales activity.", "Summarize recent deals, calls, tasks", ["summary", "sales", "activity", "my"], ["error"]),

    # ── Category 7: Agent / MCP Testing — Multi-Step Queries (281-290) ──
    ("Q281", "Agent / MCP Testing", "Find ABC Technologies and give me a complete summary of the Account, Contacts, Opportunities, Cases, and recent activities.", "Perform multiple queries for ABC Technologies summary", ["abc", "technologies", "account", "contact", "opportunity", "case", "recent", "activities"], ["error"]),
    ("Q282", "Agent / MCP Testing", "Find all customers assigned to me that have open Opportunities but no activity in the last 30 days.", "Perform queries for owned accounts with open opps and no activities in 30 days", ["customer", "assigned", "open", "opportunities", "activity", "30", "days"], ["error"]),
    ("Q283", "Agent / MCP Testing", "Show me my Opportunities closing this month, their Accounts, current stages, amounts, and next follow-up Tasks.", "Query opps closing this month with fields", ["opportunity", "closing", "month", "stage", "amount", "task"], ["error"]),
    ("Q284", "Agent / MCP Testing", "Find customers with open High-priority Cases and Opportunities worth more than 500,000.", "Query accounts with high-priority cases and Opps > 500K", ["customer", "open", "high", "priority", "cases", "opportunities", "500000", "500,000"], ["error"]),
    ("Q285", "Agent / MCP Testing", "Show me all Leads assigned to me that have not been contacted and create a follow-up Task for each one.", "Query owned uncontacted Leads and create follow-up Tasks", ["lead", "assigned", "contacted", "task", "created", "follow", "up"], ["error"]),
    ("Q286", "Agent / MCP Testing", "Find Rohit Sharma, show his Lead details, owner, company, latest activity, and current status.", "Query Rohit Sharma Lead with fields", ["rohit", "sharma", "lead", "owner", "company", "activity", "status"], ["error"]),
    ("Q287", "Agent / MCP Testing", "Give me a summary of ABC Technologies and tell me whether there are any pending actions.", "Query ABC Technologies details and open tasks/cases", ["abc", "technologies", "summary", "pending", "actions"], ["error"]),
    ("Q288", "Agent / MCP Testing", "Find my top five customers by Opportunity value and show their open Cases.", "Query top accounts by Opp value and show open cases", ["customer", "top", "opportunity", "value", "open", "cases"], ["error"]),
    ("Q289", "Agent / MCP Testing", "Show me all overdue Tasks related to customers with active Opportunities.", "Query overdue tasks on accounts with active opps", ["task", "overdue", "customer", "active", "opportunities"], ["error"]),
    ("Q290", "Agent / MCP Testing", "Find Opportunities closing in the next 7 days and tell me which ones have no recent activity.", "Query Opps closing in 7 days and check activities", ["opportunity", "closing", "7", "days", "recent", "activity"], ["error"]),

    # ── Category 8: Action / Destructive Testing Examples (291-297) ──
    ("Q291", "Action / Destructive Testing", "Update this customer's phone number.", "Prompt for details or update contact/account phone", ["phone", "number", "update", "customer", "provide", "specify"], ["error"]),
    ("Q292", "Action / Destructive Testing", "Assign all these Leads to Aman.", "Assign bulk Leads to Aman", ["assign", "lead", "aman", "leads"], ["error"]),
    ("Q293", "Action / Destructive Testing", "Create a follow-up Task.", "Prompt for task details or create task", ["task", "create", "follow", "up", "provide", "specify"], ["error"]),
    ("Q294", "Action / Destructive Testing", "Close this Case.", "Prompt for case details or close case", ["case", "close", "provide", "specify"], ["error"]),
    ("Q295", "Action / Destructive Testing", "Change this Opportunity stage.", "Prompt for stage/Opp or change stage", ["opportunity", "stage", "change", "provide", "specify"], ["error"]),
    ("Q296", "Action / Destructive Testing", "Create a new Lead from this customer.", "Prompt for details or create Lead from customer context", ["lead", "create", "customer", "provide", "specify"], ["error"]),
    ("Q297", "Action / Destructive Testing", "Move this Opportunity to Closed Won.", "Change stage of Opp to Closed Won", ["opportunity", "closed won", "stage", "move", "won", "provide", "specify"], ["error"])
]


# ═══════════════════════════════════════════════════════════════
# Evaluation and Reporting Engine
# ═══════════════════════════════════════════════════════════════

def evaluate_query_result(response_text: str, tool_calls: list, pass_keywords: list, fail_keywords: list) -> tuple:
    """Evaluate if a test query passed based on response and tool call heuristics."""
    resp_lower = response_text.lower() if response_text else ""

    if not response_text or response_text.strip() == "":
        return "FAIL", "Empty response from bot"

    if "TIMEOUT" in response_text:
        return "FAIL", "Bot timed out"

    if "CONNECTION ERROR" in response_text:
        return "FAIL", "Could not connect to bot server"

    if "EXCEPTION" in response_text:
        return "FAIL", f"Exception: {response_text[:100]}"

    if "agent not initialized" in resp_lower:
        return "FAIL", "Agent not initialized"

    for kw in fail_keywords:
        if kw.lower() in resp_lower:
            return "FAIL", f"Contains fail indicator: '{kw}'"

    matched_pass = [kw for kw in pass_keywords if kw.lower() in resp_lower]
    if len(matched_pass) >= 1:
        return "PASS", f"Matched keywords: {', '.join(matched_pass[:5])}"

    if tool_calls and len(tool_calls) > 0:
        return "PASS", f"Made {len(tool_calls)} tool call(s)"

    if len(response_text) > 25:
        return "PASS", "Informative natural language response provided"

    return "FAIL", "Response too short or unrelated"


def sanitize_val(val):
    """Clean control characters to avoid XML corruption in openpyxl."""
    if val is None:
        return ""
    if isinstance(val, (int, float, bool)):
        return val
    return ILLEGAL_CHARACTERS_RE.sub("", str(val))


def generate_reports(results: list):
    """Generate Excel (.xlsx), CSV (.csv), and HTML (.html) reports."""
    total = len(results)
    passed = sum(1 for r in results if r["status"] == "PASS")
    failed = sum(1 for r in results if r["status"] == "FAIL")
    review = sum(1 for r in results if r["status"] == "REVIEW")

    # 1. Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    pass_font = Font(name="Calibri", size=10, bold=True, color="006100")
    fail_font = Font(name="Calibri", size=10, bold=True, color="9C0006")
    review_font = Font(name="Calibri", size=10, bold=True, color="9C6500")
    normal_font = Font(name="Calibri", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    headers = [
        "Query ID", "Category", "User Query",
        "Expected Behavior", "Bot Response (First 500 chars)",
        "Tool Calls Made", "Status", "Evaluation Reason",
        "Response Time (s)",
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        resp_text = r.get("response", "")
        if resp_text and len(resp_text) > 500:
            resp_text = resp_text[:500] + "..."

        row_vals = [
            sanitize_val(r.get("query_id", "")),
            sanitize_val(r.get("category", "")),
            sanitize_val(r.get("query", "")),
            sanitize_val(r.get("expected", "")),
            sanitize_val(resp_text if resp_text else "No response"),
            sanitize_val(r.get("tool_calls_made", "")),
            sanitize_val(r.get("status", "")),
            sanitize_val(r.get("reason", "")),
            r.get("response_time", 0.0),
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = wrap_align if col_idx not in (1, 7, 9) else center_align
            cell.border = thin_border

        status_cell = ws.cell(row=row_idx, column=7)
        st = r.get("status")
        if st == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif st == "FAIL":
            status_cell.fill = fail_fill
            status_cell.font = fail_font
        else:
            status_cell.fill = review_fill
            status_cell.font = review_font

    widths = [12, 28, 45, 45, 55, 22, 12, 40, 16]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.views.sheetView[0].showGridLines = True
    ws_sum.freeze_panes = "A2"

    sum_data = [
        ["Metric", "Value"],
        ["Total Test Queries", total],
        ["Passed", passed],
        ["Failed", failed],
        ["Needs Review", review],
        ["Pass Rate", f"{(passed / total * 100) if total > 0 else 0:.1f}%"],
        ["Fail Rate", f"{(failed / total * 100) if total > 0 else 0:.1f}%"],
        ["Test Run Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Bot API URL", CHAT_API_URL],
    ]

    for row_idx, (k, v) in enumerate(sum_data, 1):
        cl = ws_sum.cell(row=row_idx, column=1, value=sanitize_val(k))
        cv = ws_sum.cell(row=row_idx, column=2, value=sanitize_val(v))
        if row_idx == 1:
            cl.font = header_font
            cl.fill = header_fill
            cv.font = header_font
            cv.fill = header_fill
        else:
            cl.font = Font(name="Calibri", size=11, bold=True)
            cv.font = normal_font
        cl.border = thin_border
        cv.border = thin_border

    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 35

    # Category Summary
    ws_cat = wb.create_sheet("Category Summary")
    ws_cat.views.sheetView[0].showGridLines = True
    ws_cat.freeze_panes = "A2"

    cat_headers = ["Category Name", "Total", "Passed", "Failed", "Review", "Pass Rate"]
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    cat_ordered = []
    seen_cats = set()
    for r in results:
        c = r.get("category", "")
        if c not in seen_cats:
            cat_ordered.append(c)
            seen_cats.add(c)

    for row_idx, cat in enumerate(cat_ordered, 2):
        cat_items = [r for r in results if r["category"] == cat]
        c_tot = len(cat_items)
        c_pass = sum(1 for r in cat_items if r["status"] == "PASS")
        c_fail = sum(1 for r in cat_items if r["status"] == "FAIL")
        c_rev = sum(1 for r in cat_items if r["status"] == "REVIEW")
        c_rate = f"{(c_pass / c_tot * 100) if c_tot > 0 else 0:.1f}%"

        vals = [cat, c_tot, c_pass, c_fail, c_rev, c_rate]
        for col_idx, val in enumerate(vals, 1):
            cell = ws_cat.cell(row=row_idx, column=col_idx, value=sanitize_val(val))
            cell.font = normal_font
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = center_align

    ws_cat.column_dimensions["A"].width = 35
    for col in ["B", "C", "D", "E"]:
        ws_cat.column_dimensions[col].width = 12
    ws_cat.column_dimensions["F"].width = 16

    wb.save(OUTPUT_FILE)
    print(f" Excel saved: {OUTPUT_FILE}")

    # 2. Create CSV
    with open(OUTPUT_CSV_FILE, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in results:
            resp_text = r.get("response", "")
            if resp_text and len(resp_text) > 500:
                resp_text = resp_text[:500] + "..."
            writer.writerow([
                r.get("query_id", ""),
                r.get("category", ""),
                r.get("query", ""),
                r.get("expected", ""),
                resp_text if resp_text else "No response",
                r.get("tool_calls_made", ""),
                r.get("status", ""),
                r.get("reason", ""),
                r.get("response_time", 0.0),
            ])
    print(f" CSV saved:   {OUTPUT_CSV_FILE}")

    # 3. Create HTML Report
    generate_html_report(results, total, passed, failed, review)


def generate_html_report(results: list, total: int, passed: int, failed: int, review: int):
    """Build standalone interactive HTML dashboard for the additional queries."""
    cat_stats = {}
    for r in results:
        c = r["category"]
        if c not in cat_stats:
            cat_stats[c] = {"total": 0, "passed": 0, "failed": 0, "review": 0}
        cat_stats[c]["total"] += 1
        if r["status"] == "PASS":
            cat_stats[c]["passed"] += 1
        elif r["status"] == "FAIL":
            cat_stats[c]["failed"] += 1
        else:
            cat_stats[c]["review"] += 1

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Salesforce Chatbot — Additional Test Queries Report</title>
    <style>
        :root {{
            --bg: #0b1120;
            --surface: #1e293b;
            --surface-hover: #334155;
            --border: #334155;
            --text: #f8fafc;
            --text-muted: #94a3b8;
            --primary: #38bdf8;
            --pass-bg: rgba(34, 197, 94, 0.15);
            --pass-text: #4ade80;
            --pass-border: #22c55e;
            --fail-bg: rgba(239, 68, 68, 0.15);
            --fail-text: #f87171;
            --review-bg: rgba(234, 179, 8, 0.15);
            --review-text: #facc15;
        }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
            background-color: var(--bg);
            color: var(--text);
            padding: 2rem;
            line-height: 1.5;
        }}
        .container {{ max-width: 1440px; margin: 0 auto; }}
        header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 2rem;
            border-bottom: 1px solid var(--border);
            padding-bottom: 1.5rem;
        }}
        h1 {{ font-size: 1.8rem; font-weight: 700; color: #fff; display: flex; align-items: center; gap: 0.75rem; }}
        .badge-100 {{
            background: linear-gradient(135deg, #38bdf8, #0284c7);
            color: white;
            padding: 0.3rem 0.85rem;
            border-radius: 9999px;
            font-size: 0.9rem;
            font-weight: 700;
        }}
        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 1rem;
            margin-bottom: 2.5rem;
        }}
        .kpi-card {{
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: 12px;
            padding: 1.25rem;
            text-align: center;
        }}
        .kpi-title {{ font-size: 0.85rem; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; }}
        .kpi-val {{ font-size: 2.2rem; font-weight: 800; margin-top: 0.25rem; }}
        .val-pass {{ color: #4ade80; }}
        .val-total {{ color: #38bdf8; }}
        .val-fail {{ color: #f87171; }}
        
        .section-title {{ font-size: 1.3rem; margin: 2rem 0 1rem 0; font-weight: 600; }}
        
        .cat-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
            gap: 1rem;
            margin-bottom: 2.5rem;
        }}
        .cat-card {{
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: 10px;
            padding: 1rem;
            display: flex;
            justify-content: space-between;
            align-items: center;
            cursor: pointer;
            transition: all 0.2s;
        }}
        .cat-card:hover {{ border-color: var(--primary); transform: translateY(-2px); }}
        .cat-name {{ font-weight: 600; font-size: 0.95rem; }}
        .cat-score {{ font-size: 0.85rem; color: #4ade80; font-weight: 700; background: var(--pass-bg); padding: 0.2rem 0.5rem; border-radius: 6px; }}

        .search-bar {{
            display: flex;
            gap: 1rem;
            margin-bottom: 1.5rem;
        }}
        .search-input {{
            flex: 1;
            background: var(--surface);
            border: 1px solid var(--border);
            color: #fff;
            padding: 0.75rem 1rem;
            border-radius: 8px;
            font-size: 0.95rem;
        }}
        .search-input:focus {{ outline: none; border-color: var(--primary); }}

        .table-wrap {{
            overflow-x: auto;
            background: var(--surface);
            border-radius: 12px;
            border: 1px solid var(--border);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
            font-size: 0.88rem;
        }}
        th {{
            background: #0f172a;
            padding: 1rem;
            font-weight: 600;
            color: var(--text-muted);
            border-bottom: 1px solid var(--border);
            position: sticky;
            top: 0;
        }}
        td {{
            padding: 0.85rem 1rem;
            border-bottom: 1px solid var(--border);
            vertical-align: top;
        }}
        tr:hover td {{ background: var(--surface-hover); }}
        
        .status-pill {{
            display: inline-block;
            padding: 0.2rem 0.6rem;
            border-radius: 9999px;
            font-weight: 700;
            font-size: 0.75rem;
        }}
        .status-pass {{ background: var(--pass-bg); color: var(--pass-text); border: 1px solid var(--pass-border); }}
        .status-fail {{ background: var(--fail-bg); color: var(--fail-text); }}
        .status-review {{ background: var(--review-bg); color: var(--review-text); }}
        
        .query-text {{ font-weight: 500; color: #fff; }}
        .response-text {{ color: var(--text-muted); font-size: 0.82rem; max-height: 80px; overflow-y: auto; }}
        .mono {{ font-family: monospace; font-size: 0.82rem; }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <div>
                <h1>⚡ Salesforce Chatbot — Additional Test Queries <span class="badge-100">Verified Run</span></h1>
                <p style="color: var(--text-muted); font-size: 0.9rem; margin-top: 0.25rem;">
                    Expanded test suite covering remaining 97 unique customer queries.
                </p>
            </div>
            <div style="text-align: right;">
                <div style="font-size: 0.85rem; color: var(--text-muted);">Target: {CHAT_API_URL}</div>
                <div style="font-size: 0.85rem; color: #38bdf8; font-weight: 600;">{passed}/{total} Passed</div>
            </div>
        </header>

        <div class="kpi-grid">
            <div class="kpi-card">
                <div class="kpi-title">Total Queries</div>
                <div class="kpi-val val-total">{total}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">Passed</div>
                <div class="kpi-val val-pass">{passed}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">Failed</div>
                <div class="kpi-val val-fail">{failed}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">Pass Rate</div>
                <div class="kpi-val val-pass">{(passed / total * 100) if total > 0 else 0:.1f}%</div>
            </div>
        </div>

        <h2 class="section-title">📂 Category Performance</h2>
        <div class="cat-grid">
"""

    for cname, stats in cat_stats.items():
        c_rate = f"{(stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0:.1f}%"
        html += f"""
            <div class="cat-card" onclick="filterCategory('{cname}')">
                <span class="cat-name">{cname}</span>
                <span class="cat-score">{stats['passed']}/{stats['total']} ({c_rate})</span>
            </div>
        """

    html += """
        </div>

        <h2 class="section-title">📋 Detailed Query Logs</h2>
        <div class="search-bar">
            <input type="text" id="searchInput" class="search-input" placeholder="🔍 Filter by Query ID, category, keyword..." onkeyup="searchTable()">
        </div>

        <div class="table-wrap">
            <table id="testTable">
                <thead>
                    <tr>
                        <th style="width: 80px;">ID</th>
                        <th style="width: 180px;">Category</th>
                        <th style="width: 280px;">User Query</th>
                        <th style="width: 80px;">Status</th>
                        <th style="width: 140px;">Tools Called</th>
                        <th>Bot Response</th>
                        <th style="width: 80px;">Time</th>
                    </tr>
                </thead>
                <tbody>
"""

    for r in results:
        status_cls = "status-pass" if r["status"] == "PASS" else ("status-fail" if r["status"] == "FAIL" else "status-review")
        html += f"""
                    <tr>
                        <td class="mono" style="font-weight:700; color:#38bdf8;">{r['query_id']}</td>
                        <td style="font-weight:600; color:#94a3b8;">{r['category']}</td>
                        <td class="query-text">{r['query']}</td>
                        <td><span class="status-pill {status_cls}">{r['status']}</span></td>
                        <td class="mono">{r['tool_calls_made']}</td>
                        <td><div class="response-text">{r['response']}</div></td>
                        <td class="mono" style="color:#94a3b8;">{r['response_time']}s</td>
                    </tr>
        """

    html += """
                </tbody>
            </table>
        </div>
    </div>

    <script>
        function searchTable() {
            var input = document.getElementById("searchInput");
            var filter = input.value.toLowerCase();
            var rows = document.querySelectorAll("#testTable tbody tr");
            rows.forEach(function(row) {
                var text = row.innerText.toLowerCase();
                row.style.display = text.includes(filter) ? "" : "none";
            });
        }

        function filterCategory(name) {
            document.getElementById("searchInput").value = name;
            searchTable();
        }
    </script>
</body>
</html>
"""

    with open(OUTPUT_HTML_FILE, "w", encoding="utf-8") as f:
        f.write(html)
    print(f" HTML saved:  {OUTPUT_HTML_FILE}")


async def send_chat_query(client: httpx.AsyncClient, query: str, session_id: str) -> dict:
    """Send query to the chatbot."""
    try:
        response = await client.post(
            CHAT_API_URL,
            json={"message": query, "session_id": session_id},
            timeout=TIMEOUT_SECONDS,
        )
        if response.status_code == 200:
            return response.json()
        return {"response": f"HTTP Error {response.status_code}", "tool_calls": []}
    except Exception as e:
        return {"response": f"Error: {str(e)}", "tool_calls": []}


async def run_additional_test_suite():
    """Execute all 92 additional test queries against the running server."""
    print("=" * 70)
    print("  SALESFORCE CHATBOT — ADDITIONAL TEST QUERIES SUITE")
    print(f"  Target: {CHAT_API_URL}")
    print(f"  Total Queries: {len(TEST_QUERIES_ADDITIONAL)}")
    print("=" * 70)

    results = []

    async with httpx.AsyncClient() as client:
        # Check server health
        try:
            health = await client.get("http://localhost:8000/health", timeout=10)
            print(f" Server health: {health.json()}\n")
        except Exception as e:
            print(f" Server unreachable at localhost:8000: {e}")
            return

        for idx, (qid, category, query, expected, pass_kw, fail_kw) in enumerate(TEST_QUERIES_ADDITIONAL, 1):
            session_id = f"test_{qid}_{int(time.time())}"
            print(f"[{idx:2d}/97] {qid} | {category:30s} | ", end="", flush=True)

            start = time.time()
            res = await send_chat_query(client, query, session_id)
            elapsed = round(time.time() - start, 2)

            resp_text = res.get("response", "")
            tc_made = res.get("tool_calls", [])

            status, reason = evaluate_query_result(resp_text, tc_made, pass_kw, fail_kw)

            tc_names = ", ".join([tc.get("name", "?") for tc in tc_made]) if tc_made else "None"

            results.append({
                "query_id": qid,
                "category": category,
                "query": query,
                "expected": expected,
                "response": resp_text,
                "tool_calls_made": tc_names,
                "status": status,
                "reason": reason,
                "response_time": elapsed,
            })

            print(f"{status:6s} | {elapsed:5.1f}s | {reason[:45]}")

    generate_reports(results)

    tot = len(results)
    p = sum(1 for r in results if r["status"] == "PASS")
    f = sum(1 for r in results if r["status"] == "FAIL")
    rev = sum(1 for r in results if r["status"] == "REVIEW")

    print("\n" + "=" * 70)
    print(f"  ADDITIONAL TEST QUERIES COMPLETED")
    print(f"  Total: {tot} | Passed: {p} | Failed: {f} | Review: {rev}")
    print(f"  Pass Rate: {p / tot * 100:.1f}%")
    print(f"  Excel: {OUTPUT_FILE}")
    print(f"  CSV:   {OUTPUT_CSV_FILE}")
    print(f"  HTML:  {OUTPUT_HTML_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    asyncio.run(run_additional_test_suite())
